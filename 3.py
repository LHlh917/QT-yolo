import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime, timedelta

# 打开原始 Excel 文件
filename = r"C:\Users\LH\Desktop\lithumiondata.xlsx"
wb = openpyxl.load_workbook(filename)
ws = wb.active

# 在第一列添加时间列
time_header = "Time"
ws.insert_cols(1)
ws.cell(row=1, column=1, value=time_header).alignment = Alignment(horizontal='center')
start_date = "01.01.2017-00:00"
time_format = "%d.%m.%Y-%H:%M"

for row in range(2, ws.max_row + 1):
    ws.cell(row=row, column=1, value=start_date).alignment = Alignment(horizontal='center')
    start_date = (datetime.strptime(start_date, time_format) + timedelta(minutes=1)).strftime(time_format)

# 保存新的 Excel 文件
new_filename = "lithumiondata1.xlsx"
wb.save(new_filename)
print("New Excel file created with time column added:", new_filename)
